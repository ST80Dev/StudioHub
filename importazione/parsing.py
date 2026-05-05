"""Parser Excel per le sessioni di import.

Strategia:
- Apre l'xlsx in modalità read-only con openpyxl (data_only=True per
  prendere i valori calcolati e non le formule).
- Identifica la riga di intestazione (header_row) e prende come nomi colonna
  i suoi valori non vuoti.
- Itera sulle righe successive distinguendo:
    * "riga di sezione": 1-2 celle non vuote, una contiene keyword note
      (DITTE INDIVIDUALI / SNC / SAS / SRL / ASSOCIAZIONI / FALLIMENTI / ...).
      Aggiorna il contesto corrente (tipo_soggetto, regime_contabile,
      contabilita) ma NON viene emessa.
    * "riga dato": ne fa un dict {nome_colonna: valore} e la emette con
      il contesto di sezione corrente.
- Le righe completamente vuote vengono saltate.

Il parser non fa match sulle anagrafiche: si limita a normalizzare la
struttura del foglio. Il match avviene in una fase successiva.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Iterator

from openpyxl import load_workbook

from anagrafica.models import (
    GestioneContabilita,
    RegimeContabile,
    TipoSoggetto,
)


# ---------------------------------------------------------------------------
# Riconoscimento sezioni
# ---------------------------------------------------------------------------

# Mappa keyword → tipo_soggetto. Ricerca case-insensitive su sottostringhe.
TIPO_SOGGETTO_KEYWORDS = [
    ("DITTE INDIVIDUALI", TipoSoggetto.DI),
    ("DITTA INDIVIDUALE", TipoSoggetto.DI),
    ("PROFESSIONIST", TipoSoggetto.PROFEX),
    ("ASSOCIAZIONI", TipoSoggetto.ASS),
    ("STUDI ASSOCIATI", TipoSoggetto.ASS),
    ("FALLIMENT", TipoSoggetto.FALL),
    # Le forme societarie vanno cercate separatamente perché la sigla puo'
    # comparire anche in stringhe diverse: gli ordini sotto contano.
    ("SOCIETA' DI CAPITALI", None),
    ("SOCIETA' DI PERSONE", None),
    (" SRL", TipoSoggetto.SRL),
    (" S.R.L.", TipoSoggetto.SRL),
    (" SPA", TipoSoggetto.SPA),
    (" S.P.A.", TipoSoggetto.SPA),
    (" SNC", TipoSoggetto.SNC),
    (" S.N.C.", TipoSoggetto.SNC),
    (" SAS", TipoSoggetto.SAS),
    (" S.A.S.", TipoSoggetto.SAS),
    ("PERSONE FISICHE", TipoSoggetto.PF),
]

REGIME_KEYWORDS = [
    ("ORDINARIA", RegimeContabile.ORDINARIO),
    ("ORDINARIO", RegimeContabile.ORDINARIO),
    ("SEMPLIFICATA", RegimeContabile.SEMPLIFICATO),
    ("SEMPLIFICATO", RegimeContabile.SEMPLIFICATO),
    ("FORFETTARI", RegimeContabile.FORFETTARIO),
    ("FORFAITARI", RegimeContabile.FORFETTARIO),
]

# Indica se la contabilità è interna allo studio o tenuta dal cliente.
CONTABILITA_INTERNA_HINTS = [
    "PRESSO LO STUDIO",
    "TENUTA DALLO STUDIO",
    "TENUTA STUDIO",
    "INTERNA",
]
CONTABILITA_ESTERNA_HINTS = [
    "ESTERNAMENTE",
    "TENUTA ESTERNA",
    "PRESSO TERZI",
    "ESTERNA",
]


@dataclass
class SectionContext:
    """Contesto dedotto dall'ultima riga-intestazione di sezione."""

    tipo_soggetto: str = ""
    regime_contabile: str = ""
    contabilita: str = ""
    raw_label: str = ""

    def to_dict(self) -> dict:
        return {
            "tipo_soggetto": self.tipo_soggetto,
            "regime_contabile": self.regime_contabile,
            "contabilita": self.contabilita,
            "raw_label": self.raw_label,
        }


def _detect_section(label: str) -> SectionContext | None:
    """Se `label` è riconoscibile come riga di sezione, restituisce il contesto.

    Match basato su sottostringhe maiuscole. Restituisce None se la stringa
    non contiene alcuna keyword: in quel caso non è una sezione.
    """
    if not label:
        return None
    upper = f" {label.upper()} "
    ctx = SectionContext(raw_label=label.strip())
    matched = False

    for kw, tipo in TIPO_SOGGETTO_KEYWORDS:
        if kw in upper:
            if tipo is not None:
                ctx.tipo_soggetto = tipo
            matched = True
            break

    for kw, regime in REGIME_KEYWORDS:
        if kw in upper:
            ctx.regime_contabile = regime
            matched = True
            break

    if any(h in upper for h in CONTABILITA_INTERNA_HINTS):
        ctx.contabilita = GestioneContabilita.INTERNA
        matched = True
    elif any(h in upper for h in CONTABILITA_ESTERNA_HINTS):
        ctx.contabilita = GestioneContabilita.ESTERNA
        matched = True

    return ctx if matched else None


# ---------------------------------------------------------------------------
# Parser principale
# ---------------------------------------------------------------------------

@dataclass
class ParsedRow:
    numero_riga: int
    dati: dict
    contesto_sezione: dict = field(default_factory=dict)


@dataclass
class ParseResult:
    sheet_name: str
    columns: list[str]
    rows: list[ParsedRow]
    sections_seen: list[str] = field(default_factory=list)


def _stringify(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _non_empty_cells(row: tuple) -> list[tuple[int, object]]:
    return [(i, v) for i, v in enumerate(row) if v not in (None, "")]


def parse_workbook(
    file_path: str,
    sheet_name: str = "",
    header_row: int = 1,
    max_rows: int = 5000,
) -> ParseResult:
    """Parsa un file xlsx restituendo struttura colonne + righe + contesto.

    `header_row` è 1-based.
    """
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    columns: list[str] = []
    rows: list[ParsedRow] = []
    sections_seen: list[str] = []
    context = SectionContext()

    for excel_row_idx, raw_row in enumerate(
        ws.iter_rows(values_only=True), start=1
    ):
        if excel_row_idx > max_rows:
            break

        if excel_row_idx < header_row:
            continue

        if excel_row_idx == header_row:
            columns = [_stringify(v) for v in raw_row]
            continue

        non_empty = _non_empty_cells(raw_row)
        if not non_empty:
            continue

        # Riga di sezione: 1-2 celle valorizzate, almeno una con keyword nota.
        if len(non_empty) <= 2:
            label = " ".join(_stringify(v) for _, v in non_empty)
            ctx = _detect_section(label)
            if ctx is not None:
                # I campi non riconosciuti dall'attuale label vengono ereditati
                # dal contesto precedente (tipico: sezione "ORDINARIA TENUTA
                # ESTERNAMENTE" non ripete "DITTE INDIVIDUALI").
                context = SectionContext(
                    tipo_soggetto=ctx.tipo_soggetto or context.tipo_soggetto,
                    regime_contabile=ctx.regime_contabile or context.regime_contabile,
                    contabilita=ctx.contabilita or context.contabilita,
                    raw_label=ctx.raw_label,
                )
                sections_seen.append(ctx.raw_label)
                continue

        # Riga dato: dict colonna→valore (solo per colonne con header).
        dati: dict = {}
        for idx, value in enumerate(raw_row):
            col_name = columns[idx] if idx < len(columns) else ""
            if not col_name:
                continue
            dati[col_name] = _stringify(value)

        if not any(dati.values()):
            continue

        rows.append(
            ParsedRow(
                numero_riga=excel_row_idx,
                dati=dati,
                contesto_sezione=context.to_dict(),
            )
        )

    wb.close()
    return ParseResult(
        sheet_name=ws.title,
        columns=[c for c in columns if c],
        rows=rows,
        sections_seen=sections_seen,
    )


# ---------------------------------------------------------------------------
# Autodetect mapping colonne (suggerimento, non vincolante)
# ---------------------------------------------------------------------------

# Pattern header → campo target su Anagrafica (o "extra:<chiave>").
# Usato come default: l'utente può sovrascriverlo nello step di mapping.
HEADER_PATTERNS: list[tuple[re.Pattern, str]] = [
    (re.compile(r"^\s*COD[\.\s]*CLI", re.I), "codice_cli"),
    (re.compile(r"^\s*COD[\.\s]*MULTI", re.I), "codice_multi"),
    (re.compile(r"^\s*COD[\.\s]*GSTU", re.I), "codice_gstudio"),
    (re.compile(r"^\s*COD[\.\s]*INTERNO", re.I), "codice_interno"),
    (re.compile(r"^\s*P[\.\s]*IVA", re.I), "partita_iva"),
    (re.compile(r"^\s*PARTITA\s+IVA", re.I), "partita_iva"),
    (re.compile(r"^\s*C[\.\s]*F[\.\s]*$", re.I), "codice_fiscale"),
    (re.compile(r"^\s*CODICE\s+FISCALE", re.I), "codice_fiscale"),
    (re.compile(r"DENOMINAZ|RAG\.?\s*SOC|NOMINATIVO|CLIENTE", re.I), "denominazione"),
    (re.compile(r"^\s*EMAIL", re.I), "email"),
    (re.compile(r"^\s*PEC", re.I), "extra:pec"),
    (re.compile(r"^\s*INIZIO\s+ATT", re.I), "data_inizio_mandato"),
    (re.compile(r"^\s*CES[/\s]*ACQ", re.I), "extra:ces_acq"),
    (re.compile(r"^\s*GRUP", re.I), "extra:gruppo"),
    (re.compile(r"REGIME", re.I), "regime_contabile"),
    (re.compile(r"PERIOD.*IVA", re.I), "periodicita_iva"),
]


def autodetect_mapping(columns: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for col in columns:
        for pattern, target in HEADER_PATTERNS:
            if pattern.search(col):
                mapping[col] = target
                break
    return mapping
